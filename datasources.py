from openpyxl import load_workbook
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.datasets import make_classification
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score


class DataSet:
    @staticmethod
    def get_data(filepath):
        wb = load_workbook(filepath)
        ws = wb.active

        input_data, output_data = [], []

        for row in ws.iter_rows(min_row=2, values_only=True):
            input_data.append(row[:(len(row) - 1)])
            output_data.append(row[(len(row) - 1)])

        return ((input_data[:(int(len(input_data) * 70 / 100))], output_data[:(int(len(output_data) * 70 / 100))]),
                (input_data[-(int(len(input_data) * 30 / 100)):], output_data[-(int(len(output_data) * 30 / 100)):]))

    @staticmethod
    def getRandomClassificationData():
        # Шаг 1: Загрузка и подготовка данных
        # В этом примере мы будем использовать синтетические данные с помощью функции make_classification из sklearn.datasets
        X, y = make_classification(n_samples=10000, n_features=10, n_classes=2)

        # Шаг 2: Разделение данных на обучающий и тестовый наборы
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2)

        # Шаг 3: Преобразование данных в массивы numpy (если они еще не являются таковыми)
        X_train = np.array(X_train)
        X_test = np.array(X_test)
        y_train = np.array(y_train)
        y_test = np.array(y_test)

        # y_train = np.where(y_train == 0, -1, 1)
        # y_test = np.where(y_test == 0, -1, 1)

        # Шаг 4: Создание модели и обучение ее на обучающих данных
        model = LogisticRegression()
        model.fit(X_train, y_train)

        # Шаг 5: Прогнозирование на тестовых данных
        predictions = model.predict(X_test)

        # Оценка точности модели
        accuracy = accuracy_score(y_test, predictions)
        # print("X: ", X_train)
        # print("y: ", y_train)
        print("Accuracy:", accuracy)
        
        return X_train, X_test, y_train, y_test


